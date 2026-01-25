# -*- coding: utf-8 -*-
import openpyxl
import json
import sys

# Set output encoding
sys.stdout.reconfigure(encoding='utf-8')

filepath = r'e:\Download\发光武器材料进度追踪表 1.52.xlsx'
wb = openpyxl.load_workbook(filepath)

print("Sheet names:")
for name in wb.sheetnames:
    print(f"  - {name}")

# Read each sheet's first few rows to understand structure
for sheet_name in wb.sheetnames:
    print(f"\n=== {sheet_name} ===")
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
        # Filter out empty cells for readability
        non_empty = [str(c) if c is not None else '' for c in row[:15]]
        print(f"Row {row_idx}: {non_empty}")
